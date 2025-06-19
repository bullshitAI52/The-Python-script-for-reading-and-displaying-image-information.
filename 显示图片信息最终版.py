import os
from PIL import Image
import tkinter as tk
from tkinter import filedialog, Listbox, messagebox, Frame
from tkinter import ttk # 用于 Progressbar, ttk.Button, ttk.Label, ttk.Scrollbar, ttk.Style, ttk.Labelframe
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from io import BytesIO
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

# --- Helper Functions ---

def _get_image_info_and_thumbnail(file_path):
    """
    获取单个图片文件的详细信息并生成缩略图。
    由 ThreadPoolExecutor 在后台调用。
    """
    info_dict = {'file_path': file_path, 'error': None, 'thumbnail_bytes': None}
    try:
        with Image.open(file_path) as img:
            width, height = img.size
            dpi = img.info.get('dpi', (72, 72))
            dpi_x = dpi[0] if dpi[0] > 0 else 72
            dpi_y = dpi[1] if dpi[1] > 0 else 72

            width_cm = (width / dpi_x) * 2.54
            height_cm = (height / dpi_y) * 2.54
            color_mode = img.mode
            img_format = img.format
            file_size_bytes = os.path.getsize(file_path)

            info_dict.update({
                'pixel_size': (width, height),
                'physical_size': (round(width_cm, 2), round(height_cm, 2)),
                'dpi': (dpi_x, dpi_y),
                'color_mode': color_mode,
                'format': img_format,
                'file_size': file_size_bytes,
            })

            # 生成缩略图 (PNG格式，可考虑JPEG以减小体积，但需注意透明度问题)
            try:
                thumb_copy = img.copy()
                thumb_copy.thumbnail((80, 80)) # 缩略图尺寸
                img_byte_arr = BytesIO()
                thumb_copy.save(img_byte_arr, format='PNG') # 'PNG' or 'JPEG'
                img_byte_arr.seek(0)
                info_dict['thumbnail_bytes'] = img_byte_arr
            except Exception as e_thumb:
                # print(f"无法为 {os.path.basename(file_path)} 创建缩略图: {e_thumb}") # 调试时可取消注释
                pass # thumbnail_bytes 保持为 None

    except Exception as e_main:
        # print(f"无法获取 {file_path} 的图片信息: {e_main}") # 调试时可取消注释
        info_dict['error'] = str(e_main)
    
    return os.path.basename(file_path), info_dict


def format_file_size(size_bytes):
    if size_bytes is None: return "N/A"
    if size_bytes >= 1024 * 1024:
        return f"{size_bytes / (1024 * 1024):.2f} MB"
    else:
        return f"{size_bytes / 1024:.2f} KB"


class ImageToolApp:
    def __init__(self, root_window):
        self.root = root_window
        self.root.title("图片信息与处理工具 v1.4 (优化版)")
        self.root.geometry("850x650") # 稍微增大窗口适应内容

        self.folder_path = None
        self.cached_image_info = {} # 键为文件名 (basename), 值为包含信息和缩略图的字典

        self._setup_ui()
        self.set_buttons_state(tk.DISABLED)
        self.select_button.config(state=tk.NORMAL)

    def _setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="10 10 10 10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        list_labelframe = ttk.Labelframe(main_frame, text="图片信息列表", padding="5 5 5 5")
        list_labelframe.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        self.listbox = Listbox(list_labelframe, width=100, height=15, activestyle='dotbox', relief=tk.GROOVE, borderwidth=2)
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        list_scrollbar_y = ttk.Scrollbar(list_labelframe, orient=tk.VERTICAL, command=self.listbox.yview)
        list_scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.listbox.config(yscrollcommand=list_scrollbar_y.set)

        list_scrollbar_x = ttk.Scrollbar(list_labelframe, orient=tk.HORIZONTAL, command=self.listbox.xview)
        list_scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.listbox.config(xscrollcommand=list_scrollbar_x.set)

        button_frame = ttk.Frame(main_frame, padding="5 0 5 0")
        button_frame.pack(fill=tk.X)

        self.select_button = ttk.Button(button_frame, text="选择图片文件夹", command=self.select_folder_and_load_info)
        self.select_button.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)

        self.rename_button = ttk.Button(button_frame, text="批量重命名图片", command=self.start_rename_task)
        self.rename_button.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        
        self.export_button = ttk.Button(button_frame, text="导出信息到Excel", command=self.start_export_task)
        self.export_button.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)

        status_frame = ttk.Frame(main_frame, padding="5 5 5 5")
        status_frame.pack(fill=tk.X)
        self.status_label = ttk.Label(status_frame, text="请选择一个包含图片的文件夹。")
        self.status_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.progress_bar = ttk.Progressbar(status_frame, orient=tk.HORIZONTAL, length=200, mode='determinate')
        self.progress_bar.pack(side=tk.RIGHT, padx=5)

    def set_buttons_state(self, state):
        self.rename_button.config(state=state)
        self.export_button.config(state=state)

    def update_status(self, message, is_error=False, is_warning=False):
        self.status_label.config(text=message)
        if is_error:
            self.status_label.config(foreground="red")
        elif is_warning:
            self.status_label.config(foreground="orange")
        else:
            self.status_label.config(foreground="black")
        self.root.update_idletasks()

    def update_progress(self, value):
        self.progress_bar['value'] = value
        self.root.update_idletasks()

    def select_folder_and_load_info(self):
        new_folder_path = filedialog.askdirectory()
        if new_folder_path:
            current_op_message = f"正在从文件夹加载图片信息: {new_folder_path}"
            if self.folder_path == new_folder_path:
                current_op_message = f"正在刷新文件夹内容: {new_folder_path}"
            
            self.folder_path = new_folder_path
            self.update_status(current_op_message)
            self.listbox.delete(0, tk.END)
            self.cached_image_info = {}
            self.update_progress(0)
            
            self.set_buttons_state(tk.DISABLED)
            self.select_button.config(state=tk.DISABLED)

            thread = threading.Thread(target=self._load_info_background_task, daemon=True)
            thread.start()
        else:
            self.update_status("文件夹选择已取消。", is_warning=True)

    def _load_info_background_task(self):
        try:
            image_files_basenames = [
                f for f in os.listdir(self.folder_path) 
                if os.path.isfile(os.path.join(self.folder_path, f)) and
                f.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff'))
            ]
            image_files_basenames.sort()
            
            temp_info_cache = {}
            total_files = len(image_files_basenames)
            processed_files_count = 0

            if total_files == 0:
                self.root.after(0, self.update_status, f"文件夹 '{os.path.basename(self.folder_path)}' 中没有找到支持的图片文件。", True, True)
                self.root.after(0, self.update_progress, 100)
                self.root.after(0, lambda: self.select_button.config(state=tk.NORMAL))
                self.root.after(0, lambda: self.set_buttons_state(tk.DISABLED))
                return

            # 使用线程池并行处理图片信息获取和缩略图生成
            # 调整 max_workers，对于I/O密集型和少量CPU操作的任务，5-10个线程通常不错
            with ThreadPoolExecutor(max_workers=min(8, os.cpu_count() + 4 if os.cpu_count() else 8)) as executor:
                # 创建future到文件名的映射
                future_to_filename = {
                    executor.submit(_get_image_info_and_thumbnail, os.path.join(self.folder_path, fname)): fname 
                    for fname in image_files_basenames
                }

                for future in as_completed(future_to_filename):
                    # filename_key = future_to_filename[future] # Basename
                    try:
                        filename_key, info_data = future.result() # _get_image_info_and_thumbnail returns (basename, data)
                        if info_data and not info_data.get('error'):
                            temp_info_cache[filename_key] = info_data
                        elif info_data and info_data.get('error'):
                            # print(f"Error processing {filename_key}: {info_data['error']}") # 调试信息
                            temp_info_cache[filename_key] = None # 标记处理失败
                        else:
                            temp_info_cache[filename_key] = None
                    except Exception as exc:
                        # filename_key_exc = future_to_filename[future]
                        # print(f'{filename_key_exc} generated an exception: {exc}') # 调试信息
                        # temp_info_cache[filename_key_exc] = None # 标记处理失败
                        pass # 已在_get_image_info_and_thumbnail内部处理部分异常

                    processed_files_count += 1
                    progress_percentage = (processed_files_count / total_files) * 100
                    self.root.after(0, self.update_progress, progress_percentage)

            self.cached_image_info = temp_info_cache
            self.root.after(0, self.display_image_info_from_cache)
            self.root.after(0, self.update_status, f"成功加载 {len(self.cached_image_info)} 张图片的信息。")
            self.root.after(0, lambda: self.select_button.config(state=tk.NORMAL))
            self.root.after(0, lambda: self.set_buttons_state(tk.NORMAL if self.cached_image_info else tk.DISABLED))

        except Exception as e:
            self.root.after(0, self.update_status, f"加载图片信息时发生错误: {e}", True)
            self.root.after(0, lambda: self.select_button.config(state=tk.NORMAL))
            self.root.after(0, lambda: self.set_buttons_state(tk.DISABLED))

    def display_image_info_from_cache(self):
        self.listbox.delete(0, tk.END)
        if not self.folder_path:
            self.update_status("请先选择一个文件夹。", is_warning=True)
            self.listbox.insert(tk.END, "  请选择文件夹以显示图片信息。")
            return

        if not self.cached_image_info:
            self.update_status(f"文件夹 '{os.path.basename(self.folder_path)}' 中没有图片信息可显示。", is_warning=True)
            self.listbox.insert(tk.END, "  没有图片信息可显示。")
            return

        display_lines = []
        for file_name, info in sorted(self.cached_image_info.items()): # 按文件名排序显示
            if info and not info.get('error'):
                line = (
                    f"{file_name}: "
                    f"{info['pixel_size'][0]}x{info['pixel_size'][1]}px, "
                    f"{info['physical_size'][0]:.2f}x{info['physical_size'][1]:.2f}cm, "
                    f"DPI:{info['dpi'][0]}x{info['dpi'][1]}, "
                    f"模式:{info['color_mode']}, "
                    f"格式:{info['format']}, "
                    f"大小:{format_file_size(info['file_size'])}"
                )
            else:
                line = f"{file_name}: 无法获取信息或处理失败"
            display_lines.append("  " + line)
        
        for line in display_lines:
            self.listbox.insert(tk.END, line)
        self.update_progress(100)

    def start_rename_task(self):
        if not self.folder_path:
            messagebox.showwarning("操作警告", "请先选择包含图片的文件夹。")
            return
        if not self.cached_image_info:
             messagebox.showinfo("提示", "当前文件夹未加载图片信息，无法执行重命名。")
             return

        if messagebox.askyesno("确认操作", "确定要根据图片的物理尺寸批量重命名选定文件夹中的图片吗？\n此操作通常不可逆，请谨慎操作！"):
            self.update_status("开始批量重命名图片...")
            self.update_progress(0)
            self.set_buttons_state(tk.DISABLED)
            self.select_button.config(state=tk.DISABLED)
            
            thread = threading.Thread(target=self._rename_images_background_task,
                                      args=(lambda p: self.root.after(0, self.update_progress, p),
                                            self._rename_completion_callback,
                                            lambda err: self.root.after(0, self._task_error_callback, err, "批量重命名")),
                                      daemon=True)
            thread.start()

    def _rename_images_background_task(self, progress_callback, completion_callback, error_callback):
        try:
            # 获取当前文件夹内的实际文件列表，以防缓存与实际不一致（尽管加载后通常一致）
            current_files_in_folder = {f for f in os.listdir(self.folder_path) if os.path.isfile(os.path.join(self.folder_path, f))}
            
            # 从缓存中筛选出当前文件夹中实际存在的、且有有效信息的图片进行重命名
            infos_to_rename = {}
            for fname, finfo in self.cached_image_info.items():
                if fname in current_files_in_folder and finfo and not finfo.get('error') and finfo.get('physical_size'):
                    infos_to_rename[fname] = finfo
            
            existing_names_on_disk = current_files_in_folder.copy() # 用于冲突检测
            renamed_count = 0
            total_images_to_process = len(infos_to_rename)
            rename_log = []

            if total_images_to_process == 0:
                if completion_callback:
                    completion_callback("没有找到符合重命名条件的图片文件。", rename_log)
                return

            for file_original_name, info in infos_to_rename.items():
                original_file_path = os.path.join(self.folder_path, file_original_name)
                physical_size = info['physical_size']
                original_ext = os.path.splitext(file_original_name)[1].lower()
                
                new_name_base = f"{int(physical_size[0])}x{int(physical_size[1])}_cm"
                new_full_name = new_name_base + original_ext
                
                temp_existing_names_for_check = existing_names_on_disk.copy()
                if file_original_name in temp_existing_names_for_check:
                    temp_existing_names_for_check.remove(file_original_name)

                count = 0
                current_new_name_attempt = new_full_name
                while current_new_name_attempt in temp_existing_names_for_check:
                    count += 1
                    current_new_name_attempt = f"{new_name_base}_{count}{original_ext}"
                new_full_name = current_new_name_attempt
                
                if new_full_name != file_original_name:
                    new_file_path = os.path.join(self.folder_path, new_full_name)
                    try:
                        os.rename(original_file_path, new_file_path)
                        rename_log.append(f"成功: '{file_original_name}' -> '{new_full_name}'")
                        existing_names_on_disk.remove(file_original_name)
                        existing_names_on_disk.add(new_full_name)
                    except Exception as e_rename:
                        rename_log.append(f"失败: 重命名 '{file_original_name}' 时出错 - {e_rename}")
                else:
                    rename_log.append(f"跳过: '{file_original_name}' 无需重命名。")
            
                renamed_count += 1
                if progress_callback:
                    progress_callback(renamed_count / total_images_to_process * 100)
        
            if completion_callback:
                completion_callback("批量重命名操作完成。", rename_log)
        except Exception as e:
            if error_callback:
                error_callback(f"批量重命名过程中发生错误: {e}")

    def _rename_completion_callback(self, message, rename_log=None):
        self.update_status(message)
        log_summary_for_mb = ""
        if rename_log:
            log_summary_lines = rename_log[:10]
            log_summary_for_mb = "\n".join(log_summary_lines)
            if len(rename_log) > 10:
                log_summary_for_mb += f"\n...等共 {len(rename_log)} 条记录。"
            
            full_message = f"{message}\n\n部分日志:\n{log_summary_for_mb}"
            if messagebox.askyesno("保存日志", f"{full_message}\n\n是否要将完整的重命名日志保存到文件？"):
                log_save_path = filedialog.asksaveasfilename(
                    defaultextension=".txt",
                    filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
                    title="保存重命名日志",
                    initialdir=self.folder_path if self.folder_path else os.path.expanduser("~")
                )
                if log_save_path:
                    try:
                        with open(log_save_path, 'w', encoding='utf-8') as f:
                            f.write(f"重命名操作结果: {message}\n\n详细日志:\n")
                            for entry in rename_log:
                                f.write(entry + "\n")
                        messagebox.showinfo("日志已保存", f"完整的重命名日志已保存到: {log_save_path}")
                    except Exception as e_save:
                        messagebox.showerror("保存失败", f"无法保存日志文件: {e_save}")
            else:
                 messagebox.showinfo("重命名完成", f"{message}\n\n部分日志:\n{log_summary_for_mb}")
        else:
            messagebox.showinfo("重命名完成", message)

        # 重命名后，强制刷新列表 (会重新启用按钮)
        self.update_status("重命名完成，正在刷新图片列表...", is_warning=True)
        if self.folder_path: # 确保folder_path仍然有效
            self.select_folder_and_load_info() # 这会处理UI状态和重新加载
        else: # 万一 folder_path 没了
            self.select_button.config(state=tk.NORMAL)
            self.set_buttons_state(tk.DISABLED)


    def start_export_task(self):
        if not self.cached_image_info:
            messagebox.showwarning("操作警告", "没有图片信息可导出。请先选择文件夹并加载信息。")
            return

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="选择Excel文件保存位置",
            initialdir=self.folder_path if self.folder_path else os.path.expanduser("~"),
            initialfile="图片信息导出.xlsx"
        )

        if not save_path:
            self.update_status("导出操作已取消 (未选择保存路径)。", is_warning=True)
            return 

        self.update_status("开始导出数据到Excel...")
        self.update_progress(0)
        self.set_buttons_state(tk.DISABLED)
        self.select_button.config(state=tk.DISABLED)

        thread = threading.Thread(target=self._save_to_excel_background_task,
                                  args=(save_path,
                                        lambda p: self.root.after(0, self.update_progress, p),
                                        self._export_completion_callback,
                                        lambda err: self.root.after(0, self._task_error_callback, err, "导出到Excel")),
                                  daemon=True)
        thread.start()

    def _save_to_excel_background_task(self, save_path, progress_callback, completion_callback, error_callback):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Image Info"
            headers = ['文件名称', '像素尺寸 (宽x高)', '物理长度 (cm)', '物理高度 (cm)', 'DPI (X x Y)', '文件大小', '图片预览']
            ws.append(headers)

            # 使用缓存的数据进行导出，筛选掉处理失败或无信息的条目
            valid_infos_to_export = {fname: finfo for fname, finfo in self.cached_image_info.items() if finfo and not finfo.get('error')}
            total_images = len(valid_infos_to_export)
            processed_count = 0

            for file_basename, info in sorted(valid_infos_to_export.items()): # 按文件名排序写入
                ws.append([
                    file_basename,
                    f"{info['pixel_size'][0]}x{info['pixel_size'][1]}" if info.get('pixel_size') else "N/A",
                    f"{info['physical_size'][0]:.2f}" if info.get('physical_size') else "N/A",
                    f"{info['physical_size'][1]:.2f}" if info.get('physical_size') else "N/A",
                    f"{info['dpi'][0]}x{info['dpi'][1]}" if info.get('dpi') else "N/A",
                    format_file_size(info.get('file_size')),
                    None 
                ])

                current_row = ws.max_row
                thumbnail_bytes = info.get('thumbnail_bytes')
                if thumbnail_bytes:
                    try:
                        thumbnail_bytes.seek(0) # 重置指针
                        excel_img = ExcelImage(thumbnail_bytes)
                        ws.add_image(excel_img, f'G{current_row}')
                        ws.row_dimensions[current_row].height = 65
                    except Exception as e_img_add:
                        # print(f"无法为 {file_basename} 插入Excel图片预览: {e_img_add}") # 调试信息
                        ws.cell(row=current_row, column=headers.index('图片预览') + 1).value = "预览加载失败"
                else:
                    ws.cell(row=current_row, column=headers.index('图片预览') + 1).value = "无预览"
                
                processed_count += 1
                if progress_callback:
                    progress_callback(processed_count / total_images * 100 if total_images > 0 else 100)

            for i, column_cells in enumerate(ws.columns):
                column_letter = get_column_letter(i + 1)
                if column_letter == 'G':
                    ws.column_dimensions[column_letter].width = 15
                    continue
                
                max_length = 0
                for cell in column_cells:
                    if cell.value is not None:
                        try:
                            cell_str_len = len(str(cell.value))
                            if cell_str_len > max_length:
                                max_length = cell_str_len
                        except: pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

            wb.save(save_path)
            if completion_callback:
                completion_callback(f"数据已成功保存至: {save_path}")
        except Exception as e:
            if error_callback:
                error_callback(f"导出Excel时发生严重错误: {e}")

    def _export_completion_callback(self, message, is_warning=False):
        self.update_status(message, is_warning=is_warning)
        self.select_button.config(state=tk.NORMAL)
        self.set_buttons_state(tk.NORMAL if self.cached_image_info else tk.DISABLED)
        if is_warning: messagebox.showwarning("导出提示", message)
        else: messagebox.showinfo("导出完成", message)

    def _task_error_callback(self, error_message, task_name):
        self.update_status(f"{task_name}过程中失败: {error_message}", is_error=True)
        self.select_button.config(state=tk.NORMAL)
        self.set_buttons_state(tk.NORMAL if self.cached_image_info else tk.DISABLED)
        self.update_progress(0) # 或100，表示任务结束
        messagebox.showerror(f"{task_name}错误", error_message)


if __name__ == "__main__":
    root = tk.Tk()
    style = ttk.Style(root)
    available_themes = style.theme_names()
    # 尝试使用更现代的主题，如果存在的话
    if 'vista' in available_themes: style.theme_use('vista')
    elif 'clam' in available_themes: style.theme_use('clam')
    elif 'alt' in available_themes: style.theme_use('alt')
    # 在某些系统上 'default' 可能是最好的选择，或者特定于OS的主题如 'aqua' (macOS)
    app = ImageToolApp(root)
    root.mainloop()